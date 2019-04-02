import json
import os
import requests
import time
import csv

from configparser import SafeConfigParser
from io import open as iopen
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import styles

from contextlib import closing

class MorningstarDownloader:

    def __init__(self, output_dir):
        self.output_dir = output_dir
        if not os.path.isdir(self.output_dir):
            os.mkdir(self.output_dir)

    def is_float(self, x):
        try:
            a = float(x)
        except ValueError:
            return False
        else:
            return True

    def is_int(self, x):
        try:
            int(x)
        except ValueError:
            return False
        else:
            return True

    def morningstar_key_ratios(self, ticker, order='desc'):
        "http://financials.morningstar.com/ajax/exportKR2CSV.html?t=FB"
        url = 'http://financials.morningstar.com/ajax/exportKR2CSV.html?t={0}&order={1}'.format(ticker, order)
        data = None
        with requests.Session() as s:
            s.headers.update({
                                 'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36'})
            download = s.get(url)
            decoded_content = download.content.decode('utf-8')
            cr = csv.reader(decoded_content.splitlines(), delimiter=',')
            data = list(cr)
        return data

    def morningstar_financials(self, ticker, reportType, period, order, columnYear, number):
        """
        reportType: is = Income Statement, cf = Cash Flow, bs = Balance Sheet
        period: 12 for annual reporting, 3 for quarterly reporting
        dataType: this doesn't seem to change and is always A
        order: asc or desc (ascending or descending)
        columnYear: 5 or 10 are the only two values supported
        number: The units of the response data. 1 = None 2 = Thousands 3 = Millions 4 = Billions
        """
        # headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
        # response = requests.get(url, headers=headers)
        base_url = 'http://financials.morningstar.com/ajax/ReportProcess4CSV.html'
        query = '?t={ticker}&reportType={reportType}&period={period}&dataType=A&order={order}&columnYear={columnYear}&number={number}'\
            .format(ticker=ticker, reportType=reportType, period=period, order=order, columnYear=columnYear, number=number)
        query.format(ticker=ticker, reportType=reportType, period=period, order=order, columnYear=columnYear, number=number)
        request_url = '{base_url}{query}'.format(base_url=base_url, query=query)
        data = None
        with requests.Session() as s:
            s.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36'})
            download = s.get(request_url)
            decoded_content = download.content.decode('utf-8')
            cr = csv.reader(decoded_content.splitlines(), delimiter=',')
            data = list(cr)
        return data

    def to_csv(self, data, filename):
        file_path = os.path.join(self.output_dir, filename)
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(data)


    def to_excel(self, data, filename):
        wb = Workbook()
        ws = wb.active
        # ws.column_dimensions['A'].width = 50
        row_index = 0
        for row in data:
            for j in range(len(row)):
                if self.is_int(row[j]):
                    row[j] = int(row[j])
                elif self.is_float(row[j]):
                    row[j] = float(row[j])
            ws.append(row)
            title_lower_case = None
            if len(row) > 0 and row[0]:
                title_lower_case = row[0].lower().strip()
            # bold styling
            if title_lower_case == 'operating expenses' or title_lower_case == 'earnings per share' \
                    or title_lower_case == 'weighted average shares outstanding' \
                    or title_lower_case == 'total operating expenses':
                ws.cell(row=row_index+1, column=1).font = styles.Font(bold=True)
            # indentation
            if title_lower_case == 'other operating expenses' or title_lower_case == 'total operating expenses' \
                    or title_lower_case == 'basic' or title_lower_case == 'diluted':
                ws.cell(row=row_index + 1, column=1).value = '    {0}'.format(ws.cell(row=row_index + 1, column=1).value)
            row_index += 1
        abs_file = os.path.join(self.output_dir, filename)
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 5)
            ws.column_dimensions[column].width = adjusted_width
        wb.save(abs_file)

    def get_tickers_csv(self, filename):
        tickers = []
        first_row = True
        with open(filename) as f:
            r = csv.reader(f, delimiter=',')
            for row in r:
                if first_row:
                    first_row = False
                else:
                    tickers.append(row[0])
        return tickers

    def download_tickers(self, tickers):
        count = 0
        for ticker in tickers:
            self.download(ticker)
            count += 1
            print('downloaded {ticker}, {count} out of {total}'.format(ticker=ticker, count=count,
                                                                      total=len(tickers)))
            time.sleep(1)

    def download(self, ticker):
        between_reports_wait_secs = 5
        # income statement
        retry_wait_secs = 5
        years = 5
        time.sleep(between_reports_wait_secs)
        income_data = self.morningstar_financials(ticker=ticker, reportType='is', period=12, order='desc', columnYear=years, number=3)
        if len(income_data) > 0:
            self.to_excel(data=income_data, filename='{ticker}_income.xlsx'.format(ticker=ticker))
        else:
            print('No income data downloaded for ticker {ticker}, retrying after {wait_secs} seconds...'.format(ticker=ticker, wait_secs = retry_wait_secs))
            time.sleep(retry_wait_secs)
            income_data = self.morningstar_financials(ticker=ticker, reportType='is', period=12, order='desc',
                                                      columnYear=years, number=3)
            if len(income_data) > 0:
                print('Success when retrying the income data download for ticker {ticker}'.format(ticker=ticker))
                self.to_excel(data=income_data, filename='{ticker}_income.xlsx'.format(ticker=ticker))
            else:
                print('Not able to download income data after try for ticker {ticker}, skipping this ticker'.format(ticker=ticker))

        time.sleep(between_reports_wait_secs)
        # balance sheet
        balance_data = self.morningstar_financials(ticker=ticker, reportType='bs', period=12, order='desc', columnYear=years, number=3)
        if len(balance_data) > 0:
            self.to_excel(data=balance_data, filename='{ticker}_balance.xlsx'.format(ticker=ticker))
        else:
            print(
                'No balance sheet data downloaded for ticker {ticker}, retrying after {wait_secs} seconds...'.format(ticker=ticker,
                                                                                                          wait_secs=retry_wait_secs))
            time.sleep(retry_wait_secs)
            balance_data = self.morningstar_financials(ticker=ticker, reportType='bs', period=12, order='desc',
                                                       columnYear=years, number=3)
            if len(balance_data) > 0:
                print('Success when retrying the balance sheet data download for ticker {ticker}'.format(ticker=ticker))
                self.to_excel(data=balance_data, filename='{ticker}_balance.xlsx'.format(ticker=ticker))
            else:
                print('Not able to download balance sheet data after try for ticker {ticker}, skipping this ticker'.format(
                    ticker=ticker))
        time.sleep(between_reports_wait_secs)
        # cash flow
        cash_data = self.morningstar_financials(ticker=ticker, reportType='cf', period=12, order='desc', columnYear=years, number=3)
        if len(cash_data) > 0:
            self.to_excel(data=cash_data, filename='{ticker}_cash.xlsx'.format(ticker=ticker))
        else:
            print(
                'No cash flow data downloaded for ticker {ticker}, retrying after {wait_secs} seconds...'.format(ticker=ticker,
                                                                                                          wait_secs=retry_wait_secs))
            time.sleep(retry_wait_secs)
            cash_data = self.morningstar_financials(ticker=ticker, reportType='cf', period=12, order='desc',
                                                    columnYear=years, number=3)
            if len(cash_data) > 0:
                print('Success when retrying the cash flow data download for ticker {ticker}'.format(ticker=ticker))
                self.to_excel(data=cash_data, filename='{ticker}_cash.xlsx'.format(ticker=ticker))
            else:
                print('Not able to download cash flow data after try for ticker {ticker}, skipping this ticker'.format(
                    ticker=ticker))

