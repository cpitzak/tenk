import os
import csv

from morningstar.models import Metrics
from morningstar.models import session

from openpyxl import Workbook
from openpyxl import load_workbook

NUMBER_FORMAT = '#,##0_);[Red](#,##0)'
DECIMAL_FORMAT = '#,##0.00_);[Red](#,##0.00)'
PERCENT_FORMAT = '0.00%_);[Red](0.00%)'


class MorningstarParser:

    def __init__(self):
        return

    def read_csv(self, abs_file):
        data = {}
        dup_counts = {}
        row_count = 0
        with open(abs_file) as f:
            csv_reader = csv.reader(f, delimiter=',')
            for row in csv_reader:
                if row_count == 1:
                    is_first_cell = True
                    key = 'years'
                    for cell in row:
                        if is_first_cell:
                            data[key] = []
                            is_first_cell = False
                        else:
                            data[key].append(cell)
                if row_count >= 2:
                    is_first_cell = True
                    key = None
                    for cell in row:
                        if is_first_cell:
                            key = cell.lower().strip()
                            if key in data:
                                if key in dup_counts:
                                    dup_counts[key] += 1
                                else:
                                    dup_counts[key] = 1
                                key = "{0}{1}".format(key, dup_counts[key])
                            data[key] = []
                            is_first_cell = False
                        else:
                            value = ''
                            if len(cell) > 0:
                                value = float(cell)
                            data[key].append(value)
                row_count += 1
        return data

    def read(self, abs_file):
        wb = load_workbook(filename=abs_file)
        ws_name = wb.sheetnames[0]
        ws = wb[ws_name]
        data = {}
        dup_counts = {}
        row_count = 0
        for row in ws.iter_rows():
            if row_count == 1:
                is_first_cell = True
                key = 'years'
                for cell in row:
                    if is_first_cell:
                        data[key] = []
                        is_first_cell = False
                    else:
                        data[key].append(cell.value)
            if row_count >= 2:
                is_first_cell = True
                key = None
                for cell in row:
                    if is_first_cell:
                        key = cell.value.lower().strip()
                        if key in data:
                            if key in dup_counts:
                                dup_counts[key] += 1
                            else:
                                dup_counts[key] = 1
                            key = "{0}{1}".format(key, dup_counts[key])
                        data[key] = []
                        is_first_cell = False
                    else:
                        data[key].append(cell.value)
            row_count += 1
        return data

    def calculate(self, income_statement_data, balance_sheet_data, cash_flow_data, years):
        table_data = {}
        table_data['report_date'] = [None] * years
        table_data['gross_profit_margin'] = [None] * years
        table_data['sga_gross_profit'] = [None] * years
        table_data['rd_gross_profit'] = [None] * years
        table_data['depreciation_gross_profit'] = [None] * years
        table_data['interest_expense_operating_income'] = [None] * years
        table_data['taxes'] = [None] * years
        table_data['net_earnings'] = [None] * years
        table_data['net_earnings_by_revenue'] = [None] * years
        table_data['eps_basic'] = [None] * years
        table_data['eps_diluted'] = [None] * years
        table_data['return_on_assets'] = [None] * years
        table_data['liabilities_by_equity'] = [None] * years
        table_data['retained_earnings'] = [None] * years
        table_data['return_on_shareholder_equity'] = [None] * years
        table_data['capital_expend'] = [None] * years
        table_data['net_stock_buy_back'] = [None] * years
        for i in range(1, years + 1):
            if 'gross profit' in income_statement_data:
                gross_profit = income_statement_data['gross profit'][i]
            elif 'revenue' in income_statement_data and not 'cost of revenue' in income_statement_data:
                gross_profit = income_statement_data['revenue'][i]
            else:
                gross_profit = None
            if 'research and development' in income_statement_data:
                research_dev = income_statement_data['research and development'][i]
            else:
                research_dev = 0
            if 'interest expenses' in income_statement_data:
                interest_expenses = income_statement_data['interest expenses'][i]
            else:
                interest_expenses = 0
            if 'years' in income_statement_data and income_statement_data['years'][i]:
                table_data['report_date'][i-1] = income_statement_data['years'][i]
            if 'revenue' in income_statement_data and income_statement_data['revenue'][i]:
                if 'cost of revenue' in income_statement_data and income_statement_data['cost of revenue'][i]:
                    gross_profit_margin = (income_statement_data['revenue'][i] -
                                           income_statement_data['cost of revenue'][i]) / income_statement_data['revenue'][i]
                else:
                    gross_profit_margin = income_statement_data['revenue'][i] / income_statement_data['revenue'][i]
                table_data['gross_profit_margin'][i-1] = gross_profit_margin
            if 'sales, general and administrative' in income_statement_data and \
                    income_statement_data['sales, general and administrative'][i] and gross_profit:
                sga = income_statement_data['sales, general and administrative'][i] / gross_profit
                table_data['sga_gross_profit'][i-1] = sga
            if gross_profit:
                table_data['rd_gross_profit'][i-1] = research_dev / gross_profit

            if 'depreciation & amortization' in cash_flow_data and \
                    cash_flow_data['depreciation & amortization'][i] and gross_profit:
                dep = cash_flow_data['depreciation & amortization'][i] / gross_profit
                table_data['depreciation_gross_profit'][i-1] = dep
            if 'operating income' in income_statement_data and income_statement_data['operating income'][i]:
                interest_expense_operating_income = interest_expenses / income_statement_data['operating income'][i]
                table_data['interest_expense_operating_income'][i-1] = interest_expense_operating_income
            if 'income before taxes' in income_statement_data and \
                    'provision for income taxes' in income_statement_data and \
                    income_statement_data['income before taxes'][i] and \
                    income_statement_data['provision for income taxes'][i] and \
                    income_statement_data['provision for income taxes'][i] != 0:
                tax = income_statement_data['provision for income taxes'][i] / \
                      income_statement_data['income before taxes'][i]
                table_data['taxes'][i-1] = tax
            if 'net income' in income_statement_data and income_statement_data['net income'][i]:
                table_data['net_earnings'][i-1] = income_statement_data['net income'][i]

                if 'revenue' in income_statement_data and income_statement_data['revenue'][i]:
                    net_earnings_divide = income_statement_data['net income'][i] / income_statement_data['revenue'][i]
                    table_data['net_earnings_by_revenue'][i-1] = net_earnings_divide

            if 'basic' in income_statement_data and income_statement_data['basic']:
                table_data['eps_basic'][i-1] = income_statement_data['basic'][i]
            if 'diluted' in income_statement_data and income_statement_data['diluted']:
                table_data['eps_diluted'][i-1] = income_statement_data['diluted'][i]
            if 'total assets' in balance_sheet_data and balance_sheet_data['total assets'][i-1] and \
                    'net income' in income_statement_data and income_statement_data['net income'][i]:
                return_on_assets = income_statement_data['net income'][i] / balance_sheet_data['total assets'][i-1]
                table_data['return_on_assets'][i-1] = return_on_assets
            if 'treasury stock' in balance_sheet_data and balance_sheet_data['treasury stock'][i-1] and \
                    balance_sheet_data['treasury stock'][i-1] != 0 and 'total liabilities' in balance_sheet_data and \
                    balance_sheet_data['total liabilities'][i-1]:
                lib_shareholder = \
                    balance_sheet_data['total liabilities'][i-1] / \
                    (balance_sheet_data["total stockholders' equity"][i-1] - balance_sheet_data['treasury stock'][i-1])
                table_data['liabilities_by_equity'][i-1] = lib_shareholder
            if 'retained earnings' in balance_sheet_data and balance_sheet_data['retained earnings'][i-1]:
                retained_earnings = balance_sheet_data['retained earnings'][i-1]
                table_data['retained_earnings'][i-1] = retained_earnings
            if 'net income' in income_statement_data and income_statement_data['net income'][i] and \
                "total stockholders' equity" in balance_sheet_data and \
                    balance_sheet_data["total stockholders' equity"][i-1]:
                return_on_shareholder_equity = income_statement_data['net income'][i] / \
                                               balance_sheet_data["total stockholders' equity"][i-1]
                table_data['return_on_shareholder_equity'][i-1] = return_on_shareholder_equity
            if 'net income' in income_statement_data and income_statement_data['net income'][i] and \
                    income_statement_data['net income'][i] != 0 and 'capital expenditure' in cash_flow_data and \
                    cash_flow_data['capital expenditure'][i]:
                capital_expend = cash_flow_data['capital expenditure'][i] / income_statement_data['net income'][i]
                table_data['capital_expend'][i-1] = capital_expend
            if 'common stock repurchased' in cash_flow_data and cash_flow_data['common stock repurchased'][i] and \
                    'common stock issued' in cash_flow_data and cash_flow_data['common stock issued'][i]:
                net_stock_buy_back = cash_flow_data['common stock repurchased'][i] + \
                                     cash_flow_data['common stock issued'][i]
                table_data['net_stock_buy_back'][i-1] = net_stock_buy_back
        return table_data

    def to_sql(self, table_data, years, ticker):
        s = session()
        for i in range(years):
            metrics = Metrics(ticker=ticker, report_date=table_data['report_date'][i],
                              gross_profit_margin=table_data['gross_profit_margin'][i],
                              sga_gross_profit=table_data['sga_gross_profit'][i],
                              rd_gross_profit=table_data['rd_gross_profit'][i],
                              depreciation_gross_profit=table_data['depreciation_gross_profit'][i],
                              interest_expense_operating_income=table_data['interest_expense_operating_income'][i],
                              taxes=table_data['taxes'][i],
                              net_earnings=table_data['net_earnings'][i],
                              net_earnings_by_revenue=table_data['net_earnings_by_revenue'][i],
                              eps_basic=table_data['eps_basic'][i],
                              eps_diluted=table_data['eps_diluted'][i],
                              return_on_assets=table_data['return_on_assets'][i],
                              liabilities_by_equity=table_data['liabilities_by_equity'][i],
                              retained_earnings=table_data['retained_earnings'][i],
                              return_on_shareholder_equity=table_data['return_on_shareholder_equity'][i],
                              capital_expend=table_data['capital_expend'][i],
                              net_stock_buy_back=table_data['net_stock_buy_back'][i])
            s.add(metrics)
        s.commit()

    def to_excel(self, table_data, income_statement_data, years, filename):
        excel_data = [[], []]
        header = ['Name', 'Computation', 'What we want']
        for i in range(1, years+1):
            if 'years' in income_statement_data:
                header.append(income_statement_data['years'][i])
        excel_data[0].append(header)
        gross_profit_margin_row = ['Gross Profit Margin', '(Revenue - COGS) / Revenue', '>40% for past 10 years'] + \
            table_data['gross_profit_margin']
        sga_row = ['', 'SGA/gross profit', '<80% is okayish; <30% is great; consistent over time'] + \
            table_data['sga_gross_profit']
        rd_row = ['', 'R&D/gross profit', 'Depends; <30% seems good'] + table_data['rd_gross_profit']
        dep_row = ['', 'Depreciation/gross profit', '<10%, but depends on industry'] + \
                  table_data['depreciation_gross_profit']
        interest_row = ['', 'Interest expenses/operating income', 'Depends on industry; <15% is good'] + \
            table_data['interest_expense_operating_income']
        tax_row = ['', 'Income tax/pretax operating income', '~35%; anything else is a red flag'] + table_data['taxes']
        net_earnings_row = ['Net earnings', 'Net earnings', 'Upward'] + table_data['net_earnings']
        net_earnings_divide_row = ['', 'Net earnings / total revenue', '> 20%; <20% but >10% could be treasure'] + \
            table_data['net_earnings_by_revenue']
        eps_row = ['', 'EPS Diluted', 'Consistent and upward'] + table_data['eps_diluted']
        return_on_assets_row = ['Return on assets', 'Net income / total assets', 'Low'] + table_data['return_on_assets']
        lib_shareholder_row = ['', 'Total liabilities / (shareholders equity - treasury stock)', '< 0.8'] + \
            table_data['liabilities_by_equity']
        retained_earnings_row = ['', 'Retained earnings', 'Growing'] + table_data['retained_earnings']
        return_on_shareholder_equity_row = ['Return on shareholders equity', 'Net earnings /shareholders equity', 'High'] + \
            table_data['return_on_shareholder_equity']
        capital_expend_row = ['', 'Capital expenditures / net earnings', '< 50%'] + table_data['capital_expend']
        net_stock_buy_back_row = ['Net stock buyback', '-stock issued - stock repurchased',
                                  'Lots is a good sign; none is not bad; stock issuance may or may not be bad'] + \
            table_data['net_stock_buy_back']

        excel_data[0].append(gross_profit_margin_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(sga_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(rd_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(dep_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(interest_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(tax_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(net_earnings_row)
        excel_data[1].append(NUMBER_FORMAT)
        excel_data[0].append(net_earnings_divide_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(eps_row)
        excel_data[1].append(DECIMAL_FORMAT)
        excel_data[0].append(return_on_assets_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(lib_shareholder_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(retained_earnings_row)
        excel_data[1].append(NUMBER_FORMAT)
        excel_data[0].append(return_on_shareholder_equity_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(capital_expend_row)
        excel_data[1].append(PERCENT_FORMAT)
        excel_data[0].append(net_stock_buy_back_row)
        excel_data[1].append(NUMBER_FORMAT)

        wb = Workbook()
        ws = wb.active
        for row in excel_data[0]:
            ws.append(row)
        for index in range(len(excel_data[0]) - 1):
            for i in range(4, 14):
                ws.cell(row=index + 2, column=i).number_format = excel_data[1][index]
        wb.save(filename)

    def process_morningstar_data(self, income_statement, balance_sheet,
                                 cash_flow, output_dir, years, ticker):
        income_statement_data = self.read_csv(income_statement)
        balance_sheet_data = self.read_csv(balance_sheet)
        cash_flow_data = self.read_csv(cash_flow)
        table_data = self.calculate(income_statement_data, balance_sheet_data, cash_flow_data, years)
        abs_output_file = os.path.join(output_dir, '{ticker}.xlsx'.format(ticker=ticker))
        if not os.path.isdir(output_dir):
            os.mkdir(output_dir)
        self.to_excel(table_data=table_data, income_statement_data=income_statement_data, years=years, filename=abs_output_file)
        self.to_sql(table_data=table_data, years=5, ticker=ticker)
